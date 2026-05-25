"""
Strict DraftSheets ingestion for draft market data.

Run with:
    python -m superagent.data.ingest_draft_sheets --file path/to/sheet.xlsx --source draftsheetsv6 --season 2025
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from superagent.canonical_resolution import normalize_player_name
from superagent.db import SessionLocal
from superagent.models import (
    CanonicalPlayerAlias,
    DraftImportReview,
    DraftMarketImport,
    DraftPlayerMarket,
    DraftSourceRank,
    ExternalPlayerMapping,
    PlayerSeason,
)


class DraftIngestionError(ValueError):
    """Raised when draft market input fails strict validation."""


REQUIRED_COLUMNS = {"player", "team", "pos"}
RANK_SOURCE_COLUMNS = {
    "espn": "ESPN",
    "sleeper": "Sleeper",
    "nfl": "NFL",
    "rtsports": "RTSports",
    "fantrax": "Fantrax",
    "yahoo": "Yahoo",
    "fpros ecr": "Fpros ECR",
    "fantasypros": "FantasyPros",
    "2qb": "2QB",
}


def _clean_header(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalized_header(value: Any) -> str:
    return _clean_header(value).lower()


def _safe_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    return text


def _parse_float(value: Any, field: str, row_number: int, required: bool = False) -> float | None:
    if value is None or value == "":
        if required:
            raise DraftIngestionError(f"Row {row_number}: missing required numeric field '{field}'")
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            if required:
                raise DraftIngestionError(f"Row {row_number}: missing required numeric field '{field}'")
            return None
        if value.upper() in {"#N/A", "#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "N/A", "NA"}:
            if required:
                raise DraftIngestionError(f"Row {row_number}: invalid numeric value for '{field}': {value!r}")
            return None
        value = value.replace(",", "")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise DraftIngestionError(
            f"Row {row_number}: invalid numeric value for '{field}': {value!r}"
        ) from exc


def _parse_int(value: Any, field: str, row_number: int, required: bool = False) -> int | None:
    parsed = _parse_float(value, field, row_number, required=required)
    if parsed is None:
        return None
    if parsed != int(parsed):
        raise DraftIngestionError(f"Row {row_number}: '{field}' must be an integer, got {value!r}")
    return int(parsed)


def _parse_optional_int_lenient(value: Any) -> int | None:
    """Parse optional workbook context fields, treating malformed cells as absent."""
    try:
        return _parse_int(value, "optional_integer", 0)
    except DraftIngestionError:
        return None


def _split_position_rank(pos_value: Any) -> tuple[str | None, int | None]:
    text = _safe_text(pos_value)
    if not text:
        return None, None
    match = re.match(r"^([A-Za-z]+)\s*(\d+)?$", text)
    if not match:
        return text.upper(), None
    position = match.group(1).upper()
    rank = int(match.group(2)) if match.group(2) else None
    return position, rank


def _row_to_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    row = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        row[header] = values[index] if index < len(values) else None
    return row


def _validate_headers(headers: list[str]) -> None:
    normalized = {_normalized_header(header) for header in headers if header}
    missing = sorted(REQUIRED_COLUMNS - normalized)
    if missing:
        raise DraftIngestionError(f"Missing required column(s): {', '.join(missing)}")


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = [_clean_header(header) for header in (reader.fieldnames or [])]
        _validate_headers(headers)
        rows = []
        for row in reader:
            rows.append({_clean_header(key): value for key, value in row.items() if key})
        return rows


def _read_xlsx_rows(path: Path, sheet_name: str | None = None) -> tuple[list[dict[str, Any]], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DraftIngestionError(
            "XLSX ingestion requires openpyxl. Install dependencies with `pip install -r requirements.txt`."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    selected_sheet = sheet_name or ("DATA" if "DATA" in workbook.sheetnames else workbook.sheetnames[0])
    if selected_sheet not in workbook.sheetnames:
        raise DraftIngestionError(f"Sheet '{selected_sheet}' not found in {path.name}")

    worksheet = workbook[selected_sheet]
    header_row_index = None
    headers: list[str] = []
    rows: list[dict[str, Any]] = []
    for row_index, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        candidate_headers = [_clean_header(value) for value in values]
        normalized = {_normalized_header(header) for header in candidate_headers if header}
        if REQUIRED_COLUMNS.issubset(normalized):
            header_row_index = row_index
            headers = candidate_headers
            break

    if header_row_index is None:
        raise DraftIngestionError(
            f"Could not find a header row with required columns: {', '.join(sorted(REQUIRED_COLUMNS))}"
        )

    _validate_headers(headers)
    for values in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        row = _row_to_dict(headers, values)
        if _safe_text(row.get("Player")):
            rows.append(row)
    workbook.close()
    return rows, selected_sheet


def _load_rows(path: Path, sheet_name: str | None) -> tuple[list[dict[str, Any]], str | None]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(path), None
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_rows(path, sheet_name=sheet_name)
    raise DraftIngestionError(f"Unsupported draft market file type: {path.suffix}")


def _lookup(row: dict[str, Any], *names: str) -> Any:
    normalized = {_normalized_header(key): value for key, value in row.items()}
    for name in names:
        if _normalized_header(name) in normalized:
            return normalized[_normalized_header(name)]
    return None


def _draft_row_payload(row: dict[str, Any], row_number: int) -> dict[str, Any]:
    player_name = _safe_text(_lookup(row, "Player", "PLAYER NAME", "source_player_name"))
    team = _safe_text(_lookup(row, "Team", "Tm", "TM"))
    position, position_rank = _split_position_rank(_lookup(row, "POS", "Position"))
    if not player_name:
        raise DraftIngestionError(f"Row {row_number}: missing player name")
    if not position:
        return {
            "source_player_name": player_name,
            "team": team,
            "position": None,
            "position_rank": None,
            "bye_week": _parse_optional_int_lenient(_lookup(row, "Bye", "Bye Week")),
            "overall_rank": _parse_float(_lookup(row, "Rank", "RK"), "overall_rank", row_number),
            "adp": None,
            "ecr": None,
            "avg_rank": None,
            "best_rank": None,
            "worst_rank": None,
            "std_dev": None,
            "ecr_vs_adp": None,
            "floor": None,
            "ceiling": None,
            "value": None,
            "injury_risk": None,
            "raw_data": {key: value for key, value in row.items() if value not in (None, "")},
            "source_ranks": {},
        }
    payload = {
        "source_player_name": player_name,
        "team": team,
        "position": position,
        "position_rank": position_rank,
        "bye_week": _parse_optional_int_lenient(_lookup(row, "Bye", "Bye Week")),
        "overall_rank": _parse_float(_lookup(row, "Rank", "RK"), "overall_rank", row_number),
        "adp": _parse_float(_lookup(row, "ADP"), "adp", row_number),
        "ecr": _parse_float(_lookup(row, "Fpros ECR", "ECR"), "ecr", row_number),
        "avg_rank": _parse_float(_lookup(row, "AVG", "AVG."), "avg_rank", row_number),
        "best_rank": _parse_float(_lookup(row, "BEST"), "best_rank", row_number),
        "worst_rank": _parse_float(_lookup(row, "WORST"), "worst_rank", row_number),
        "std_dev": _parse_float(_lookup(row, "STD.DEV", "STD DEV"), "std_dev", row_number),
        "ecr_vs_adp": _parse_float(_lookup(row, "ECR VS. ADP"), "ecr_vs_adp", row_number),
        "floor": _parse_float(_lookup(row, "Floor", "LOW"), "floor", row_number),
        "ceiling": _parse_float(_lookup(row, "Ceiling", "HIGH"), "ceiling", row_number),
        "value": _parse_float(_lookup(row, "Value", "Val"), "value", row_number),
        "injury_risk": _safe_text(_lookup(row, "Injury Risk", "Risk")),
        "raw_data": {key: value for key, value in row.items() if value not in (None, "")},
        "source_ranks": {},
    }
    for normalized_source, source_name in RANK_SOURCE_COLUMNS.items():
        value = _lookup(row, source_name)
        rank = _parse_float(value, source_name, row_number)
        if rank is not None:
            payload["source_ranks"][source_name] = rank
            if source_name == "Fpros ECR" and payload["ecr"] is None:
                payload["ecr"] = rank
    return payload


def _upsert_market_row(
    db: Session,
    import_batch: DraftMarketImport,
    source: str,
    season: int,
    canonical_player_id: str,
    payload: dict[str, Any],
) -> None:
    market = (
        db.query(DraftPlayerMarket)
        .filter(
            DraftPlayerMarket.source == source,
            DraftPlayerMarket.season == season,
            DraftPlayerMarket.canonical_player_id == canonical_player_id,
        )
        .first()
    )
    if market is None:
        market = DraftPlayerMarket(
            import_id=import_batch.id,
            source=source,
            season=season,
            canonical_player_id=canonical_player_id,
            source_player_name=payload["source_player_name"],
        )
        db.add(market)
        db.flush()

    market.import_id = import_batch.id
    market.source_player_name = payload["source_player_name"]
    market.team = payload["team"]
    market.position = payload["position"]
    market.position_rank = payload["position_rank"]
    market.bye_week = payload["bye_week"]
    market.overall_rank = payload["overall_rank"]
    market.adp = payload["adp"]
    market.ecr = payload["ecr"]
    market.avg_rank = payload["avg_rank"]
    market.best_rank = payload["best_rank"]
    market.worst_rank = payload["worst_rank"]
    market.std_dev = payload["std_dev"]
    market.ecr_vs_adp = payload["ecr_vs_adp"]
    market.floor = payload["floor"]
    market.ceiling = payload["ceiling"]
    market.value = payload["value"]
    market.injury_risk = payload["injury_risk"]
    market.raw_data = json.dumps(payload["raw_data"], default=str)

    for rank_source, rank_value in payload["source_ranks"].items():
        source_rank = (
            db.query(DraftSourceRank)
            .filter(
                DraftSourceRank.market_id == market.id,
                DraftSourceRank.rank_source == rank_source,
            )
            .first()
        )
        if source_rank is None:
            source_rank = DraftSourceRank(market_id=market.id, rank_source=rank_source, rank_value=rank_value)
            db.add(source_rank)
        else:
            source_rank.rank_value = rank_value


def _build_exact_alias_index(db: Session, season: int) -> dict[str, list[dict[str, Any]]]:
    """Preload exact alias candidates once for fast draft import mapping."""
    seasons_by_player: dict[str, list[PlayerSeason]] = {}
    for player_season in db.query(PlayerSeason).filter(PlayerSeason.season == season).all():
        seasons_by_player.setdefault(player_season.canonical_player_id, []).append(player_season)

    candidates_by_alias: dict[str, list[dict[str, Any]]] = {}
    aliases = db.query(CanonicalPlayerAlias).all()
    seen = set()
    for alias in aliases:
        player = alias.canonical_player
        season_contexts = seasons_by_player.get(player.canonical_player_id, [])
        if not season_contexts:
            season_contexts = [None]
        for context in season_contexts:
            key = (
                alias.normalized_alias,
                player.canonical_player_id,
                context.position if context else None,
                context.team if context else None,
            )
            if key in seen:
                continue
            seen.add(key)
            candidates_by_alias.setdefault(alias.normalized_alias, []).append(
                {
                    "canonical_player_id": player.canonical_player_id,
                    "full_name": player.full_name,
                    "position": context.position if context else None,
                    "team": context.team if context else None,
                    "confidence": 1.0,
                    "source": "exact_alias",
                }
            )
    return candidates_by_alias


def _upsert_external_mapping(
    db: Session,
    source: str,
    season: int,
    source_player_name: str,
    canonical_player_id: str,
    confidence: float,
) -> None:
    mapping = (
        db.query(ExternalPlayerMapping)
        .filter(
            ExternalPlayerMapping.source == source,
            ExternalPlayerMapping.season == season,
            ExternalPlayerMapping.source_player_name == source_player_name,
            ExternalPlayerMapping.source_player_id.is_(None),
        )
        .first()
    )
    if mapping is None:
        mapping = ExternalPlayerMapping(
            source=source,
            season=season,
            source_player_name=source_player_name,
            source_player_id=None,
        )
        db.add(mapping)
    mapping.canonical_player_id = canonical_player_id
    mapping.confidence = confidence
    mapping.status = "auto"


def _queue_draft_review(
    db: Session,
    source: str,
    season: int,
    source_player_name: str,
    candidates: list[dict[str, Any]],
) -> None:
    db.add(
        DraftImportReview(
            source=source,
            season=season,
            source_player_name=source_player_name,
            source_player_id=None,
            candidates=json.dumps(candidates),
            status="pending",
        )
    )


def _map_draft_player_exact(
    db: Session,
    source: str,
    season: int,
    source_player_name: str,
    position: str | None,
    alias_index: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    normalized = normalize_player_name(source_player_name)
    candidates = alias_index.get(normalized, [])
    if position:
        position_matches = [
            candidate
            for candidate in candidates
            if candidate.get("position")
            and str(candidate["position"]).upper() == position.upper()
        ]
        if position_matches:
            candidates = position_matches

    unique_by_player = {}
    for candidate in candidates:
        unique_by_player.setdefault(candidate["canonical_player_id"], candidate)
    candidates = list(unique_by_player.values())

    if len(candidates) == 1:
        candidate = candidates[0]
        _upsert_external_mapping(
            db=db,
            source=source,
            season=season,
            source_player_name=source_player_name,
            canonical_player_id=candidate["canonical_player_id"],
            confidence=candidate["confidence"],
        )
        return {
            "ok": True,
            "canonical_player_id": candidate["canonical_player_id"],
            "confidence": candidate["confidence"],
            "status": "auto",
        }

    _queue_draft_review(
        db=db,
        source=source,
        season=season,
        source_player_name=source_player_name,
        candidates=candidates[:5],
    )
    return {
        "ok": False,
        "error": "needs_review" if candidates else "no_exact_alias",
        "needs_review": True,
        "candidates": candidates[:5],
    }


def ingest_draft_market_file(
    file_path: str,
    source: str,
    season: int,
    sheet_name: str | None = None,
    db: Session | None = None,
) -> dict[str, Any]:
    """Import DraftSheets-style market data, mapping rows through canonical identity."""
    path = Path(file_path)
    if not path.exists():
        raise DraftIngestionError(f"Draft market file not found: {file_path}")
    if not source.strip():
        raise DraftIngestionError("Source is required")
    if season < 2020 or season > 2030:
        raise DraftIngestionError(f"Invalid draft market season: {season}")

    rows, actual_sheet_name = _load_rows(path, sheet_name)
    if not rows:
        raise DraftIngestionError("Draft market file has no player rows")

    owns_session = db is None
    db = db or SessionLocal()
    try:
        existing = (
            db.query(DraftMarketImport)
            .filter(
                DraftMarketImport.source == source,
                DraftMarketImport.season == season,
                DraftMarketImport.file_name == path.name,
            )
            .first()
        )
        if existing is not None:
            db.delete(existing)
            db.flush()

        import_batch = DraftMarketImport(
            source=source,
            season=season,
            file_name=path.name,
            sheet_name=actual_sheet_name,
            status="completed",
            rows_seen=len(rows),
        )
        db.add(import_batch)
        db.flush()

        rows_imported = 0
        rows_needing_review = 0
        review_rows = []
        alias_index = _build_exact_alias_index(db, season)
        for index, row in enumerate(rows, start=2):
            payload = _draft_row_payload(row, index)
            mapping = _map_draft_player_exact(
                db=db,
                source=source,
                season=season,
                source_player_name=payload["source_player_name"],
                position=payload["position"],
                alias_index=alias_index,
            )
            if not mapping["ok"]:
                rows_needing_review += 1
                review_rows.append(
                    {
                        "row": index,
                        "player": payload["source_player_name"],
                        "position": payload["position"],
                        "reason": mapping.get("error"),
                    }
                )
                continue

            _upsert_market_row(
                db=db,
                import_batch=import_batch,
                source=source,
                season=season,
                canonical_player_id=mapping["canonical_player_id"],
                payload=payload,
            )
            rows_imported += 1

        import_batch.rows_imported = rows_imported
        import_batch.rows_needing_review = rows_needing_review
        if rows_needing_review:
            import_batch.status = "completed_with_review"
        db.commit()
        return {
            "ok": True,
            "import_id": import_batch.id,
            "source": source,
            "season": season,
            "file_name": path.name,
            "sheet_name": actual_sheet_name,
            "rows_seen": len(rows),
            "rows_imported": rows_imported,
            "rows_needing_review": rows_needing_review,
            "review_rows": review_rows,
        }
    except Exception:
        db.rollback()
        raise
    finally:
        if owns_session:
            db.close()


def main() -> None:
    """CLI entrypoint for strict DraftSheets import."""
    parser = argparse.ArgumentParser(description="Import strict draft market data")
    parser.add_argument("--file", "--csv", dest="file_path", required=True, help="Path to CSV/XLSX file")
    parser.add_argument("--source", required=True, help="Source id, e.g. draftsheetsv6")
    parser.add_argument("--season", required=True, type=int, help="Draft season")
    parser.add_argument("--sheet", dest="sheet_name", help="Optional XLSX sheet name. Defaults to DATA.")
    args = parser.parse_args()

    summary = ingest_draft_market_file(
        file_path=args.file_path,
        source=args.source,
        season=args.season,
        sheet_name=args.sheet_name,
    )
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
